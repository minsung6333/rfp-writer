# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["ID", "카테고리", "명칭", "정의", "세부내용", "방향/힌트", "초안", "상태"]
STATUS_TODO = "미작성"
STATUS_DRAFT = "초안완료"
STATUS_REVIEW = "검토중"
STATUS_DONE = "완료"

HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
COL_WIDTHS = [12, 20, 30, 40, 60, 30, 60, 10]


def create_excel(requirements: list[dict], output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "요구사항"

    # 헤더
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 20

    # 데이터
    for row, req in enumerate(requirements, 2):
        values = [
            req.get("id", ""),
            req.get("category", ""),
            req.get("name", ""),
            req.get("definition", ""),
            req.get("detail", ""),
            "",  # 방향/힌트
            "",  # 초안
            STATUS_TODO,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            cell.font = Font(name="Arial", size=10)

        ws.row_dimensions[row].height = 60

    wb.save(output_path)
    return output_path


def load_requirements(excel_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["요구사항"]
    requirements = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        requirements.append({
            "id": row[0] or "",
            "category": row[1] or "",
            "name": row[2] or "",
            "definition": row[3] or "",
            "detail": row[4] or "",
            "hint": row[5] or "",
            "draft": row[6] or "",
            "status": row[7] or STATUS_TODO,
        })
    return requirements


def save_requirement(excel_path: str, req_id: str, field: str, value: str):
    field_to_col = {"hint": 6, "draft": 7, "status": 8}
    col = field_to_col.get(field)
    if not col:
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["요구사항"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == req_id:
            row[col - 1].value = value
            break
    wb.save(excel_path)
