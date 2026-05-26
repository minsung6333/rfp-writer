# -*- coding: utf-8 -*-
import os, json, tempfile, io, datetime
import streamlit as st

# API 키: .env → st.secrets 순서로 탐색
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# secrets에서 API 키 로딩 (Streamlit Cloud 배포용)
if not os.getenv("OPENAI_API_KEY"):
    try:
        os.environ["OPENAI_API_KEY"] = st.secrets["OPENAI_API_KEY"]
    except Exception:
        pass

st.set_page_config(page_title="RFP 제안서 작성기", page_icon="📝", layout="wide")

# ── 비밀번호 잠금 ──
def _check_password():
    """앱 접근 비밀번호 확인. secrets에 APP_PASSWORD가 없으면 잠금 해제."""
    try:
        correct_pw = st.secrets["APP_PASSWORD"]
    except Exception:
        return True  # 비밀번호 미설정 → 잠금 없음 (로컬 개발용)
    if st.session_state.get("authenticated"):
        return True
    st.title("🔒 RFP 제안서 작성기")
    pw = st.text_input("비밀번호를 입력하세요", type="password", key="pw_input")
    if st.button("로그인", type="primary"):
        if pw == correct_pw:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("비밀번호가 틀렸습니다.")
    return False

if not _check_password():
    st.stop()

from modules.parser import (
    extract_text_by_page, find_requirement_pages, find_toc_pages,
    parse_requirements_with_llm, parse_toc_with_llm, generate_toc_from_requirements,
    parse_page_range, parse_pages_with_vision, extract_toc_with_vision,
)
from modules.slide_manager import bulk_link_reqs
from modules.slide_manager import (
    classify_slide, generate_questions,
    build_slides_from_toc, get_uncovered_reqs,
)
from modules.drafter import (
    generate_draft_stream, generate_draft_from_answers_stream,
    revise_draft_stream, review_draft,
    generate_outline, generate_section_stream, generate_section,
    generate_chapter_outlines,
)

STATUS_ICON = {"미작성": "⬜", "질문중": "💬", "초안완료": "📝", "검토중": "⚠️", "완료": "✅"}

for key, default in [
    ("requirements", []),
    ("slides", []),
    ("project_overview", {}),
    ("idea_chats", {}),
    ("selected_slide_id", None),
    ("qa_buffers", {}),
    ("debate", {
        "chapter": None,
        "status": "idle",
        "history": [],
        "round": 0,
        "max_rounds": 2,
        "pending_question": "",
        "pending_agent": "",
        "pending_doc_request": "",
        "pending_doc_queue": [],
        "ref_text_cache": "",
        "selected_reqs": [],
        "final_strategy": "",
        "user_addition": "",
        "active_ideas": [],
    }),
]:
    if key not in st.session_state:
        st.session_state[key] = default


def get_slide(slide_id: str) -> dict | None:
    for s in st.session_state.slides:
        if s["id"] == slide_id:
            return s
    return None


def update_slide(slide_id: str, **kwargs):
    for s in st.session_state.slides:
        if s["id"] == slide_id:
            s.update(kwargs)
            break


def get_req(req_id: str) -> dict | None:
    for r in st.session_state.requirements:
        if r["id"] == req_id:
            return r
    return None


# ── 사이드바 ───────────────────────────────────────────────
with st.sidebar:
    st.title("📝 RFP 제안서 작성기")
    st.divider()

    # 입력 방식 탭
    st.subheader("1. 데이터 입력")
    tab_pdf, tab_xl, tab_session = st.tabs(["📄 PDF 파싱", "📥 엑셀 로드", "💾 세션"])

    with tab_pdf:
        uploaded = st.file_uploader("PDF 파일", type=["pdf"], label_visibility="collapsed")
        if uploaded and st.button("파싱 시작", use_container_width=True, type="primary"):
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(uploaded.read())
                tmp_path = tmp.name
            with st.status("파싱 중...", expanded=True) as status:
                from modules.parser import (
                    extract_text_by_page, find_requirement_pages, find_toc_pages,
                    split_req_chunks, parse_chunk, parse_toc_with_llm,
                    generate_toc_from_requirements, extract_project_overview,
                )
                from concurrent.futures import ThreadPoolExecutor, as_completed

                st.write("📄 PDF 텍스트 추출 중...")
                pages = extract_text_by_page(tmp_path)
                st.write(f"✅ {len(pages)}페이지 추출 완료")

                req_text, is_fallback = find_requirement_pages(pages)
                toc_text = find_toc_pages(pages)
                chunks = split_req_chunks(req_text)

                st.write(f"🔍 요구사항 파싱 중... (총 {len(chunks)}개 청크, gpt-4o-mini)")
                progress = st.progress(0)
                req_status = st.empty()

                all_requirements, seen_ids = [], set()
                with ThreadPoolExecutor(max_workers=4) as executor:
                    f_toc = executor.submit(parse_toc_with_llm, toc_text)
                    futures = {executor.submit(parse_chunk, c): i for i, c in enumerate(chunks)}
                    done = 0
                    for future in as_completed(futures):
                        done += 1
                        progress.progress(done / len(chunks))
                        req_status.caption(f"청크 {done}/{len(chunks)} 완료")
                        for req in future.result():
                            rid = req.get("id", "")
                            if rid and rid not in seen_ids:
                                seen_ids.add(rid)
                                all_requirements.append(req)
                    toc = f_toc.result()

                progress.empty()
                req_status.empty()

                toc_auto = False
                if not toc and all_requirements:
                    st.write("⚠️ 목차 추출 실패 → 요구사항 기반 목차 자동 생성 중...")
                    toc = generate_toc_from_requirements(all_requirements)
                    toc_auto = True

                for r in all_requirements:
                    r.setdefault("status", "미작성")
                    r.setdefault("hint", "")
                    r.setdefault("draft", "")
                st.write(f"✅ 요구사항 {len(all_requirements)}개 / 목차 {len(toc)}개{'(자동생성)' if toc_auto else ''} 추출")

                st.write("🔗 장표-요구사항 매핑 중...")
                slides = build_slides_from_toc(toc, all_requirements)
                st.write(f"✅ 장표 {len(slides)}개 생성 완료")

                st.write("📋 사업 개요 추출 중...")
                overview = extract_project_overview(pages)
                if overview:
                    pname = overview.get("project_name", "(미상)")
                    st.write(f"✅ 사업 개요 추출: {pname}")

                st.session_state.requirements = all_requirements
                st.session_state.slides = slides
                st.session_state.project_overview = overview
                st.session_state.selected_slide_id = None
                st.session_state.qa_buffers = {}
                status.update(label="파싱 완료!", state="complete")
            os.remove(tmp_path)
            if toc_auto:
                st.warning(f"⚠️ 제안서 목차를 찾지 못해 요구사항 기반으로 목차를 자동 생성했습니다.\n요구사항 {len(all_requirements)}개, 장표 {len(slides)}개")
            elif is_fallback:
                st.warning(f"⚠️ 표준 요구사항 ID를 찾지 못해 전체 문서에서 과업을 추출했습니다.\n요구사항 {len(all_requirements)}개, 장표 {len(slides)}개")
            else:
                st.success(f"✅ 요구사항 {len(all_requirements)}개, 장표 {len(slides)}개 추출 완료")

    with tab_xl:
        st.caption("파싱 결과 엑셀을 수정 후 재업로드하면 LLM 없이 바로 로드됩니다.")
        uploaded_xl = st.file_uploader("엑셀 파일", type=["xlsx"], key="xl_upload",
                                       label_visibility="collapsed")
        if uploaded_xl and st.button("로드", use_container_width=True, type="primary"):
            try:
                import pandas as pd
                xl = pd.ExcelFile(uploaded_xl)

                df_req = pd.read_excel(xl, sheet_name="요구사항", dtype=str).fillna("")
                col_map_req = {"ID": "id", "카테고리": "category", "명칭": "name",
                               "정의": "definition", "세부내용": "detail"}
                df_req = df_req.rename(columns=col_map_req)
                loaded_reqs = df_req[["id", "category", "name", "definition", "detail"]].to_dict("records")
                for r in loaded_reqs:
                    r.setdefault("status", "미작성")
                    r.setdefault("hint", "")
                    r.setdefault("draft", "")

                loaded_slides = []
                if "장표 목차" in xl.sheet_names:
                    df_toc = pd.read_excel(xl, sheet_name="장표 목차", dtype=str).fillna("")
                    col_map_toc = {"장표 ID": "id", "챕터": "chapter", "섹션": "section",
                                   "제목": "title", "연결 요구사항": "linked_reqs_str", "상태": "status"}
                    df_toc = df_toc.rename(columns=col_map_toc)
                    for _, row in df_toc.iterrows():
                        linked = [x.strip() for x in row.get("linked_reqs_str", "").split(",") if x.strip()]
                        loaded_slides.append({
                            "id": row.get("id", ""),
                            "chapter": row.get("chapter", ""),
                            "section": row.get("section", ""),
                            "title": row.get("title", ""),
                            "linked_reqs": linked,
                            "type": "",
                            "status": row.get("status", "미작성"),
                            "questions": [],
                            "draft": "",
                            "chat_history": [],
                        })
                else:
                    st.warning("장표 목차 시트가 없습니다. 요구사항만 로드됩니다.")

                # 사업 개요 시트가 있으면 로드
                loaded_overview = {}
                if "사업 개요" in xl.sheet_names:
                    df_ov = pd.read_excel(xl, sheet_name="사업 개요", dtype=str).fillna("")
                    label_to_key = {
                        "사업명": "project_name", "추진 배경": "background",
                        "추진 필요성": "necessity", "사업 목적": "purpose",
                        "주요 목표": "goals", "사업 범위": "scope",
                        "기간·예산": "duration_budget", "핵심 정보": "key_points",
                    }
                    for _, row in df_ov.iterrows():
                        label = str(row.get("항목", "")).strip()
                        content = str(row.get("내용", "")).strip()
                        if label in label_to_key and content:
                            loaded_overview[label_to_key[label]] = content

                st.session_state.requirements = loaded_reqs
                st.session_state.slides = loaded_slides
                st.session_state.project_overview = loaded_overview
                st.session_state.selected_slide_id = None
                st.session_state.qa_buffers = {}
                ov_msg = f", 사업 개요 ✓" if loaded_overview else ""
                st.success(f"✅ 요구사항 {len(loaded_reqs)}개, 장표 {len(loaded_slides)}개{ov_msg} 로드 완료")
                st.rerun()
            except Exception as e:
                st.error(f"엑셀 로드 오류: {e}")

    with tab_session:
        st.caption("작업 중인 프로젝트를 JSON 파일로 저장/복원합니다. (전략·토론·VLLM 결과·장표 초안 전체 포함)")

        # 저장
        has_work = bool(st.session_state.get("requirements") or st.session_state.get("slides"))
        if has_work:
            session_data = {
                "version": "v2",
                "saved_at": datetime.datetime.now().isoformat(timespec="seconds"),
                "requirements": st.session_state.get("requirements", []),
                "slides": st.session_state.get("slides", []),
                "project_overview": st.session_state.get("project_overview", {}),
                "idea_chats": st.session_state.get("idea_chats", {}),
                "debate": st.session_state.get("debate", {}),
                "qa_buffers": st.session_state.get("qa_buffers", {}),
                "selected_slide_id": st.session_state.get("selected_slide_id"),
            }
            try:
                session_json = json.dumps(session_data, ensure_ascii=False, indent=2)
                fname = f"rfp_session_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                st.download_button(
                    "💾 프로젝트 저장",
                    data=session_json.encode("utf-8"),
                    file_name=fname,
                    mime="application/json",
                    use_container_width=True,
                )
                st.caption(f"크기: {len(session_json):,}자")
            except Exception as e:
                st.error(f"저장 직렬화 실패: {e}")
        else:
            st.info("저장할 작업이 없습니다. PDF 파싱 또는 엑셀 로드를 먼저 진행하세요.")

        st.divider()

        # 복원
        uploaded_session = st.file_uploader(
            "세션 JSON 파일", type=["json"], key="session_upload", label_visibility="collapsed",
        )
        if uploaded_session and st.button("📂 불러오기", use_container_width=True, type="primary"):
            try:
                data = json.loads(uploaded_session.read().decode("utf-8"))
                for k in ["requirements", "slides", "project_overview", "idea_chats", "debate", "qa_buffers", "selected_slide_id"]:
                    if k in data:
                        st.session_state[k] = data[k]
                st.success(f"✅ 복원 완료 (저장 시각: {data.get('saved_at', '-')})")
                st.rerun()
            except Exception as e:
                st.error(f"복원 실패: {e}")

    st.divider()

    # 사업 개요
    if st.session_state.get("project_overview"):
        from modules.parser import format_overview_text
        ov = st.session_state.project_overview
        pname = ov.get("project_name", "사업 개요")
        with st.expander(f"📋 {pname[:30]}", expanded=False):
            labels = {
                "project_name": "사업명",
                "background": "추진 배경",
                "necessity": "추진 필요성",
                "purpose": "사업 목적",
                "goals": "주요 목표",
                "scope": "사업 범위",
                "duration_budget": "기간·예산",
                "key_points": "핵심 정보",
            }
            for key, label in labels.items():
                cur = ov.get(key, "")
                new_v = st.text_area(
                    label, value=cur, height=70 if key != "goals" else 100,
                    key=f"overview_{key}",
                )
                if new_v != cur:
                    st.session_state.project_overview[key] = new_v
            st.caption("편집 가능 — LLM 호출 시 컨텍스트로 자동 전달됩니다.")

        st.divider()

    # 진행 현황
    if st.session_state.slides:
        total = len(st.session_state.slides)
        done = sum(1 for s in st.session_state.slides if s["status"] == "완료")
        drafted = sum(1 for s in st.session_state.slides if s["status"] in ["초안완료", "검토중"])
        st.progress(done / total if total else 0)
        st.caption(f"완료 {done} / 초안 {drafted} / 전체 {total}")

        # 미커버 요구사항 (카운트만, 자세한건 탭에서)
        uncovered = get_uncovered_reqs(st.session_state.requirements, st.session_state.slides)
        if uncovered:
            st.caption(f"⚠️ 미커버 {len(uncovered)}개 → **요구사항 목록** 탭에서 확인")


# ── 엑셀 내보내기 ──────────────────────────────────────────
def build_excel(requirements: list[dict], slides: list[dict], overview: dict = None) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    wb = Workbook()

    # 사업 개요 시트 (있을 때만, 첫 시트로)
    if overview:
        ws_ov = wb.active
        ws_ov.title = "사업 개요"
        labels = [
            ("project_name", "사업명"),
            ("background", "추진 배경"),
            ("necessity", "추진 필요성"),
            ("purpose", "사업 목적"),
            ("goals", "주요 목표"),
            ("scope", "사업 범위"),
            ("duration_budget", "기간·예산"),
            ("key_points", "핵심 정보"),
        ]
        # 헤더
        for col, h in enumerate(["항목", "내용"], 1):
            cell = ws_ov.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="C65911")
            cell.alignment = Alignment(horizontal="center")
        for row, (key, label) in enumerate(labels, 2):
            ws_ov.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws_ov.cell(row=row, column=2, value=overview.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_ov.column_dimensions["A"].width = 18
        ws_ov.column_dimensions["B"].width = 80
        for r in range(2, len(labels) + 2):
            ws_ov.row_dimensions[r].height = 60

        ws_req = wb.create_sheet("요구사항")
    else:
        ws_req = wb.active
        ws_req.title = "요구사항"
    headers_req = ["ID", "카테고리", "명칭", "정의", "세부내용"]
    for col, h in enumerate(headers_req, 1):
        cell = ws_req.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2F5496")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row, r in enumerate(requirements, 2):
        ws_req.cell(row=row, column=1, value=r.get("id", ""))
        ws_req.cell(row=row, column=2, value=r.get("category", ""))
        ws_req.cell(row=row, column=3, value=r.get("name", ""))
        ws_req.cell(row=row, column=4, value=r.get("definition", ""))
        cell = ws_req.cell(row=row, column=5, value=r.get("detail", ""))
        cell.alignment = Alignment(wrap_text=True)
    for col, width in zip(range(1, 6), [12, 16, 28, 30, 60]):
        ws_req.column_dimensions[get_column_letter(col)].width = width

    # 목차/장표 시트
    ws_toc = wb.create_sheet("장표 목차")
    headers_toc = ["장표 ID", "챕터", "섹션", "제목", "연결 요구사항", "상태"]
    for col, h in enumerate(headers_toc, 1):
        cell = ws_toc.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="375623")
        cell.alignment = Alignment(horizontal="center")
    for row, s in enumerate(slides, 2):
        ws_toc.cell(row=row, column=1, value=s.get("id", ""))
        ws_toc.cell(row=row, column=2, value=s.get("chapter", ""))
        ws_toc.cell(row=row, column=3, value=s.get("section", ""))
        ws_toc.cell(row=row, column=4, value=s.get("title", ""))
        ws_toc.cell(row=row, column=5, value=", ".join(s.get("linked_reqs", [])))
        ws_toc.cell(row=row, column=6, value=s.get("status", ""))
    for col, width in zip(range(1, 7), [12, 8, 8, 36, 28, 10]):
        ws_toc.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _vllm_with_progress(fn, label: str = "VLLM 처리"):
    """fn(progress_callback) → result, with st.progress bar."""
    bar = st.progress(0.0, text=f"{label} 준비 중...")
    def cb(done: int, total: int):
        if total > 0:
            bar.progress(min(done / total, 1.0), text=f"{label} {done}/{total} 배치 완료")
    try:
        return fn(cb)
    finally:
        bar.empty()


# ── 메인 화면 ──────────────────────────────────────────────
if not st.session_state.slides:
    st.info("👈 사이드바에서 RFP PDF를 업로드하세요.")
    st.stop()

# 파싱 결과 확인 탭
tab_strategy, tab_ideas, tab_reqs, tab_toc = st.tabs(
    ["🎯 전략 수립", "💡 기능 아이디어 도출", "📊 요구사항 목록", "🗂️ 목차 구성"]
)

with tab_reqs:
    import pandas as pd
    reqs = st.session_state.requirements
    slides = st.session_state.slides

    # 역매핑: req_id → [slide titles]
    req_to_slides = {}
    for s in slides:
        for rid in s.get("linked_reqs", []):
            req_to_slides.setdefault(rid, []).append(
                f"{s.get('chapter','')}-{s.get('section','')} {s.get('title','')}"
            )
    covered_ids = set(req_to_slides.keys())
    uncovered_reqs = [r for r in reqs if r["id"] not in covered_ids]

    # ── 커버리지 요약 ──
    col_total, col_cov, col_uncov, col_pct = st.columns(4)
    total = len(reqs)
    n_cov = total - len(uncovered_reqs)
    pct = (n_cov / total * 100) if total else 0
    col_total.metric("전체 요구사항", total)
    col_cov.metric("커버됨", n_cov)
    col_uncov.metric("미커버", len(uncovered_reqs),
                     delta=f"-{len(uncovered_reqs)}" if uncovered_reqs else None,
                     delta_color="inverse")
    col_pct.metric("커버리지", f"{pct:.1f}%")
    st.progress(pct / 100)

    # 엑셀 다운로드
    excel_bytes = build_excel(reqs, slides, st.session_state.get("project_overview"))
    if excel_bytes:
        st.download_button(
            "⬇️ 엑셀로 내보내기 (요구사항+목차)",
            data=excel_bytes,
            file_name="rfp_파싱결과.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.divider()

    # ── 카테고리별 커버리지 ──
    cat_stats = {}
    for r in reqs:
        cat = r.get("category", "(미분류)")
        cat_stats.setdefault(cat, {"total": 0, "covered": 0})
        cat_stats[cat]["total"] += 1
        if r["id"] in covered_ids:
            cat_stats[cat]["covered"] += 1
    cat_rows = [
        {
            "카테고리": cat,
            "전체": v["total"],
            "커버": v["covered"],
            "미커버": v["total"] - v["covered"],
            "커버리지": f"{v['covered']/v['total']*100:.0f}%" if v["total"] else "-",
        }
        for cat, v in sorted(cat_stats.items(), key=lambda x: -(x[1]["total"] - x[1]["covered"]))
    ]
    with st.expander("📊 카테고리별 커버리지", expanded=True):
        st.dataframe(cat_rows, use_container_width=True, hide_index=True)

    # ── 미커버 요구사항 (상세) ──
    if uncovered_reqs:
        with st.expander(f"⚠️ 미커버 요구사항 상세 ({len(uncovered_reqs)}개) — 평가 감점 위험", expanded=True):
            st.caption("아래 요구사항들은 어떤 장표에도 매핑되어 있지 않습니다. 목차 구성 탭에서 매핑을 추가하거나, 자동 재매핑을 실행하세요.")
            uncov_rows = [
                {
                    "ID": r.get("id", ""),
                    "카테고리": r.get("category", ""),
                    "명칭": r.get("name", ""),
                    "정의": r.get("definition", "") or (r.get("detail", "")[:80] + "...") if r.get("detail") else "",
                }
                for r in uncovered_reqs
            ]
            st.dataframe(uncov_rows, use_container_width=True, hide_index=True, height=min(400, 40 + 35 * len(uncov_rows)))

            if st.button("🔄 미커버 요구사항 자동 매핑 시도", use_container_width=True):
                with st.spinner("LLM이 미커버 요구사항을 적절한 장표에 매핑 중..."):
                    titles = [s.get("title", "") for s in slides]
                    # 모든 슬라이드에 대해 전체 재매핑 (단순화)
                    mapping = bulk_link_reqs(titles, reqs)
                    for s in slides:
                        new_links = mapping.get(s.get("title", ""), [])
                        # 기존 매핑 + 새로 추가된 것만 (덮어쓰지 않음)
                        existing = set(s.get("linked_reqs", []))
                        added = [r for r in new_links if r not in existing]
                        s["linked_reqs"] = list(existing) + added
                st.success("✅ 재매핑 완료")
                st.rerun()
    else:
        st.success(f"✅ 모든 요구사항({total}개)이 장표에 매핑되어 있습니다.")

    st.divider()

    # ── 전체 요구사항 표 (연결 장표 컬럼 추가) ──
    st.markdown("**전체 요구사항**")
    req_rows = [
        {
            "ID": r.get("id", ""),
            "카테고리": r.get("category", ""),
            "명칭": r.get("name", ""),
            "정의": r.get("definition", ""),
            "연결 장표": " | ".join(req_to_slides.get(r["id"], [])) or "❌ 미매핑",
            "세부내용": r.get("detail", ""),
        }
        for r in reqs
    ]
    st.dataframe(req_rows, use_container_width=True, height=500, hide_index=True)

with tab_toc:
    st.caption(f"총 {len(st.session_state.slides)}개 장표 — 표를 직접 편집하거나 VLLM으로 목차를 재추출할 수 있습니다.")

    # VLLM 목차 추출
    with st.expander("📷 이미지/VLLM으로 목차 추출 + 요구사항 자동 매핑", expanded=False):
        st.caption("RFP의 작성지침 또는 평가기준표 페이지를 지정하면 VLLM이 표를 분석해 목차와 요구사항 매핑까지 자동 생성합니다.")
        toc_v_pdf = st.file_uploader("RFP PDF", type=["pdf"], key="toc_vision_pdf")
        toc_v_pages = st.text_input(
            "목차/평가기준 페이지 (예: 30-35, 40)",
            key="toc_vision_pages",
        )
        col_v_mode, col_v_run = st.columns([1, 1])
        with col_v_mode:
            replace_mode = st.radio(
                "추출 결과 처리",
                ["기존 목차에 추가", "기존 목차 대체"],
                index=1,
                horizontal=False,
                key="toc_vision_mode",
            )
        with col_v_run:
            run_v = st.button(
                "VLLM 목차 추출 + 매핑",
                use_container_width=True,
                type="primary",
                disabled=not (toc_v_pdf and toc_v_pages.strip()),
                key="toc_vision_run",
            )
        if run_v:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(toc_v_pdf.read())
                tmp_path = tmp.name
            pages_sel = parse_page_range(toc_v_pages)
            new_slides_raw = _vllm_with_progress(
                lambda cb: extract_toc_with_vision(tmp_path, pages_sel, progress_callback=cb),
                label=f"VLLM 목차 추출 ({len(pages_sel)}p)",
            )
            os.remove(tmp_path)
            if not new_slides_raw:
                st.error("목차를 추출하지 못했습니다. 페이지 번호와 PDF 내용을 확인해주세요.")
            else:
                with st.spinner(f"{len(new_slides_raw)}개 장표 ↔ 요구사항 자동 매핑 중..."):
                    titles = [s["title"] for s in new_slides_raw]
                    mapping = bulk_link_reqs(titles, st.session_state.requirements)
                start_idx = len(st.session_state.slides) if replace_mode == "기존 목차에 추가" else 0
                new_slides = []
                for i, s in enumerate(new_slides_raw):
                    new_slides.append({
                        "id": f"slide_v_{start_idx + i:03d}",
                        "chapter": s["chapter"],
                        "section": s["section"],
                        "title": s["title"],
                        "linked_reqs": mapping.get(s["title"], []),
                        "status": "미작성",
                    })
                if replace_mode == "기존 목차 대체":
                    st.session_state.slides = new_slides
                else:
                    st.session_state.slides.extend(new_slides)
                st.success(f"✅ {len(new_slides)}개 장표 추출 및 매핑 완료")
                st.rerun()

    st.divider()

    # ── 슬라이드 편집/추가 모달 ──
    reqs_all = st.session_state.requirements
    req_map_all = {r["id"]: r for r in reqs_all}
    STATUS_OPTS = ["미작성", "초안완료", "검토중", "완료"]

    @st.dialog("슬라이드 편집", width="large")
    def edit_slide_dialog(idx=None):
        is_new = idx is None
        if is_new:
            slide = {"chapter": "", "section": "", "title": "", "linked_reqs": [], "status": "미작성"}
        else:
            slide = dict(st.session_state.slides[idx])

        col_c, col_s, col_st = st.columns([1, 1, 2])
        chapter_v = col_c.text_input("챕터", value=slide.get("chapter", ""), key=f"dlg_ch_{idx}")
        section_v = col_s.text_input("섹션", value=slide.get("section", ""), key=f"dlg_sec_{idx}")
        status_v = col_st.selectbox(
            "상태", STATUS_OPTS,
            index=STATUS_OPTS.index(slide.get("status", "미작성")) if slide.get("status") in STATUS_OPTS else 0,
            key=f"dlg_status_{idx}",
        )
        title_v = st.text_input("제목 *", value=slide.get("title", ""), key=f"dlg_title_{idx}")

        # AI 추천 (제목 기반)
        rec_key = f"dlg_rec_{idx}"
        col_ms, col_rec = st.columns([5, 1])
        with col_rec:
            if st.button("✨ AI 추천", key=f"dlg_btn_rec_{idx}", help="현재 제목으로 요구사항을 LLM이 추천", use_container_width=True):
                if title_v.strip():
                    with st.spinner("추천 중..."):
                        rec = bulk_link_reqs([title_v], reqs_all).get(title_v, [])
                    st.session_state[rec_key] = rec
                    st.rerun()
                else:
                    st.warning("제목 입력 먼저")

        default_reqs = st.session_state.get(rec_key, slide.get("linked_reqs", []))
        with col_ms:
            linked_v = st.multiselect(
                f"연결 요구사항 ({len(reqs_all)}개 중 선택)",
                options=list(req_map_all.keys()),
                default=default_reqs,
                format_func=lambda rid: f"{rid} — {req_map_all.get(rid, {}).get('name', '')[:60]}",
                key=f"dlg_reqs_{idx}",
            )

        # 선택된 요구사항 상세
        if linked_v:
            with st.expander(f"📑 선택된 요구사항 상세 ({len(linked_v)}개)", expanded=False):
                for rid in linked_v:
                    r = req_map_all.get(rid)
                    if not r:
                        continue
                    st.markdown(f"**`{rid}` {r.get('name', '')}**")
                    cap_parts = []
                    if r.get("category"): cap_parts.append(f"분류: {r['category']}")
                    if cap_parts: st.caption(" · ".join(cap_parts))
                    if r.get("definition"):
                        st.markdown(f"- 정의: {r['definition']}")
                    if r.get("detail"):
                        st.markdown(f"- 세부: {r['detail'][:400]}{'...' if len(r.get('detail',''))>400 else ''}")
                    st.divider()

        # 액션 버튼
        st.markdown("---")
        if is_new:
            col_save, col_cancel = st.columns([3, 1])
            with col_save:
                if st.button("💾 추가", use_container_width=True, type="primary", key=f"dlg_save_{idx}"):
                    if not title_v.strip():
                        st.error("제목은 필수입니다.")
                    else:
                        st.session_state.slides.append({
                            "id": f"slide_{len(st.session_state.slides):03d}",
                            "chapter": chapter_v.strip(),
                            "section": section_v.strip(),
                            "title": title_v.strip(),
                            "linked_reqs": linked_v,
                            "status": status_v,
                        })
                        st.session_state.pop(rec_key, None)
                        st.rerun(scope="app")
            with col_cancel:
                if st.button("취소", use_container_width=True, key=f"dlg_cancel_{idx}"):
                    st.session_state.pop(rec_key, None)
                    st.rerun(scope="app")
        else:
            col_save, col_cancel, col_del = st.columns([2, 1, 1])
            with col_save:
                if st.button("💾 저장", use_container_width=True, type="primary", key=f"dlg_save_{idx}"):
                    if not title_v.strip():
                        st.error("제목은 필수입니다.")
                    else:
                        existing = st.session_state.slides[idx]
                        existing.update({
                            "chapter": chapter_v.strip(),
                            "section": section_v.strip(),
                            "title": title_v.strip(),
                            "linked_reqs": linked_v,
                            "status": status_v,
                        })
                        st.session_state.pop(rec_key, None)
                        st.rerun(scope="app")
            with col_cancel:
                if st.button("취소", use_container_width=True, key=f"dlg_cancel_{idx}"):
                    st.session_state.pop(rec_key, None)
                    st.rerun(scope="app")
            with col_del:
                if st.button("🗑️ 삭제", use_container_width=True, key=f"dlg_del_{idx}"):
                    st.session_state.slides.pop(idx)
                    st.session_state.pop(rec_key, None)
                    st.rerun(scope="app")

    # 상단 액션 바
    col_add, col_remap, col_dl = st.columns([1, 1, 1])
    with col_add:
        if st.button("➕ 슬라이드 추가", use_container_width=True, type="primary"):
            edit_slide_dialog()
    with col_remap:
        if st.button("🔄 요구사항 재매핑 (전체)", use_container_width=True,
                     help="현재 목차의 제목들로 요구사항 매핑을 LLM이 다시 수행합니다 (기존 매핑 덮어씀)"):
            with st.spinner("매핑 중..."):
                titles = [s["title"] for s in st.session_state.slides]
                mapping = bulk_link_reqs(titles, reqs_all)
                for s in st.session_state.slides:
                    s["linked_reqs"] = mapping.get(s["title"], [])
            st.success("재매핑 완료")
            st.rerun()
    with col_dl:
        excel_bytes2 = build_excel(reqs_all, st.session_state.slides, st.session_state.get("project_overview"))
        if excel_bytes2:
            st.download_button(
                "⬇️ 엑셀 (개요+요구사항+목차)",
                data=excel_bytes2,
                file_name="rfp_요구사항_목차.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    st.divider()

    # ── 슬라이드 목록 (편집 버튼) ──
    if not st.session_state.slides:
        st.info("슬라이드가 없습니다. '➕ 슬라이드 추가' 또는 위 VLLM 추출을 사용하세요.")
    else:
        for idx, slide in enumerate(st.session_state.slides):
            linked = slide.get("linked_reqs", [])
            col_info, col_edit, col_del = st.columns([7, 1, 1])
            with col_info:
                ch_sec = f"{slide.get('chapter','')}-{slide.get('section','')}".strip("-")
                title = slide.get("title", "(제목 없음)")
                status = slide.get("status", "미작성")
                status_icon = {"미작성": "⬜", "초안완료": "✏️", "검토중": "🔍", "완료": "✅"}.get(status, "⬜")
                st.markdown(f"{status_icon} **{ch_sec} · {title}**")
                if linked:
                    names = []
                    for rid in linked[:4]:
                        r = req_map_all.get(rid)
                        nm = r.get("name", "")[:18] if r else ""
                        names.append(f"`{rid}` {nm}")
                    more = f" 외 {len(linked)-4}개" if len(linked) > 4 else ""
                    st.caption(f"📎 {' · '.join(names)}{more}")
                else:
                    st.caption("📎 _연결 요구사항 없음_")
            with col_edit:
                if st.button("✏️", key=f"row_edit_{idx}", use_container_width=True, help="편집"):
                    edit_slide_dialog(idx)
            with col_del:
                if st.button("🗑️", key=f"row_del_{idx}", use_container_width=True, help="삭제"):
                    st.session_state.slides.pop(idx)
                    st.rerun()

# ── 💡 기능 아이디어 도출 탭 ───────────────────────────────────────────
with tab_ideas:
    from modules.idea_chat import generate_initial_message, agentic_chat, agentic_chat_stream, summarize_idea_chat
    from modules.parser import format_overview_text as _fmt_ov, extract_text_by_page as _extract_pages, parse_pages_with_vision as _parse_vision, parse_page_range as _parse_range

    overview_text_ideas = _fmt_ov(st.session_state.get("project_overview", {}))
    reqs_idea = st.session_state.requirements
    req_map_idea = {r["id"]: r for r in reqs_idea}
    idea_chats = st.session_state.idea_chats

    col_top, col_chat = st.columns([1, 3])

    with col_top:
        st.subheader("주제 관리")

        # 새 주제 만들기
        with st.expander("➕ 새 주제", expanded=not bool(idea_chats)):
            new_topic_name = st.text_input("주제명", placeholder="예: RAG 검색 정확도", key="new_topic_name")
            new_topic_reqs = st.multiselect(
                "연결 요구사항 (선택)",
                options=[r["id"] for r in reqs_idea],
                format_func=lambda rid: f"{rid} — {req_map_idea.get(rid, {}).get('name', '')[:40]}",
                key="new_topic_reqs",
            )
            if st.button("➕ 주제 만들기", use_container_width=True, type="primary", disabled=not new_topic_name.strip()):
                tname = new_topic_name.strip()
                if tname in idea_chats:
                    st.error("이미 있는 주제명입니다.")
                else:
                    linked = [req_map_idea[rid] for rid in new_topic_reqs if rid in req_map_idea]
                    with st.spinner("AI 초기 분석 중..."):
                        initial = generate_initial_message(tname, linked, overview_text_ideas)
                    idea_chats[tname] = {
                        "linked_reqs": new_topic_reqs,
                        "messages": [{"role": "assistant", "content": initial,
                                      "search_calls": [], "citations": [], "need_doc": ""}],
                        "summary": "",
                        "use_search": True,
                    }
                    st.session_state.active_idea_topic = tname
                    st.rerun()

        # 기존 주제 목록
        st.markdown("**주제 목록**")
        if not idea_chats:
            st.caption("아직 주제가 없습니다. 위에서 새 주제를 만드세요.")
        else:
            active = st.session_state.get("active_idea_topic")
            if active not in idea_chats:
                active = list(idea_chats.keys())[0]
                st.session_state.active_idea_topic = active
            for tname, chat in idea_chats.items():
                msg_count = len(chat.get("messages", []))
                sum_icon = "✅" if chat.get("summary") else "⬜"
                is_active = tname == active
                btn_label = f"{'▶ ' if is_active else ''}{sum_icon} {tname} ({msg_count})"
                if st.button(btn_label, key=f"topic_btn_{tname}", use_container_width=True):
                    st.session_state.active_idea_topic = tname
                    st.rerun()

        # 주제 삭제
        if idea_chats and st.session_state.get("active_idea_topic"):
            active = st.session_state.active_idea_topic
            with st.expander("⚙️ 주제 설정"):
                # 연결 요구사항 편집
                cur_reqs = idea_chats[active].get("linked_reqs", [])
                new_reqs = st.multiselect(
                    "연결 요구사항",
                    options=[r["id"] for r in reqs_idea],
                    default=cur_reqs,
                    format_func=lambda rid: f"{rid} — {req_map_idea.get(rid, {}).get('name', '')[:40]}",
                    key=f"edit_reqs_{active}",
                )
                if new_reqs != cur_reqs:
                    idea_chats[active]["linked_reqs"] = new_reqs

                use_search = st.toggle(
                    "🔍 웹 검색 활성화",
                    value=idea_chats[active].get("use_search", True),
                    key=f"use_search_{active}",
                    help="끄면 LLM이 자체 지식만으로 답변 (빠름·저렴)",
                )
                idea_chats[active]["use_search"] = use_search

                if st.button("🗑️ 주제 삭제", use_container_width=True, key=f"del_topic_{active}"):
                    del idea_chats[active]
                    st.session_state.pop("active_idea_topic", None)
                    st.rerun()

    with col_chat:
        active = st.session_state.get("active_idea_topic")
        if not active or active not in idea_chats:
            st.info("👈 왼쪽에서 주제를 만들거나 선택하세요.")
        else:
            chat = idea_chats[active]
            linked_objs = [req_map_idea[rid] for rid in chat.get("linked_reqs", []) if rid in req_map_idea]
            st.subheader(f"💬 {active}")
            if linked_objs:
                st.caption("📎 " + " · ".join([f"`{r['id']}` {r.get('name','')[:25]}" for r in linked_objs[:5]]) + (f" 외 {len(linked_objs)-5}개" if len(linked_objs) > 5 else ""))

            # 채팅 history 표시
            for i, msg in enumerate(chat["messages"]):
                role = msg["role"]
                with st.chat_message(role, avatar="🤖" if role == "assistant" else "👤"):
                    content = msg.get("content", "")
                    st.markdown(content)

                    # 검색 호출 표시
                    search_calls = msg.get("search_calls", [])
                    if search_calls:
                        with st.expander(f"🔍 {len(search_calls)}회 웹 검색 수행"):
                            for sc in search_calls:
                                q = sc.get("query", "")
                                if q:
                                    st.caption(f"• {q}")

                    # 인용 표시
                    citations = msg.get("citations", [])
                    if citations:
                        with st.expander(f"📎 출처 {len(citations)}건"):
                            for c in citations:
                                title = c.get("title", "(제목 없음)")
                                url = c.get("url", "")
                                st.markdown(f"- [{title}]({url})")

                    # 문서 요청
                    need_doc = msg.get("need_doc", "")
                    doc_handled = msg.get("doc_handled", False)
                    if need_doc and not doc_handled and role == "assistant":
                        st.warning(f"📎 **문서 요청:** {need_doc}")
                        col_up, col_pg, col_skip = st.columns([3, 2, 1])
                        with col_up:
                            up_doc = st.file_uploader(
                                "파일 첨부", type=["pdf", "jpg", "jpeg", "png", "gif", "webp", "bmp"],
                                key=f"idea_doc_{active}_{i}",
                                label_visibility="collapsed",
                            )
                        with col_pg:
                            nd_pages = st.text_input(
                                "페이지", placeholder="예: 1-5, 10 (PDF만)",
                                key=f"idea_doc_pages_{active}_{i}",
                                label_visibility="collapsed",
                                disabled=not (up_doc and up_doc.name.lower().endswith(".pdf")),
                            )
                        if up_doc and st.button("문서 제출", key=f"idea_doc_submit_{active}_{i}", use_container_width=True, type="primary"):
                            from modules.parser import parse_pages_with_vision, parse_images_with_vision
                            fname = up_doc.name.lower()
                            file_bytes = up_doc.read()
                            if fname.endswith(".pdf"):
                                import tempfile, os as _os
                                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                                    tmp.write(file_bytes)
                                    tmp_path = tmp.name
                                target_pages = []
                                if nd_pages.strip():
                                    for part in nd_pages.split(","):
                                        part = part.strip()
                                        if "-" in part:
                                            s, e = part.split("-", 1)
                                            try: target_pages.extend(range(int(s), int(e) + 1))
                                            except ValueError: pass
                                        else:
                                            try: target_pages.append(int(part))
                                            except ValueError: pass
                                if not target_pages:
                                    import fitz
                                    doc_f = fitz.open(tmp_path)
                                    total_p = len(doc_f)
                                    doc_f.close()
                                    target_pages = list(range(1, min(total_p + 1, 11)))
                                with st.spinner(f"📄 VLLM 문서 분석 중... ({len(target_pages)}페이지)"):
                                    doc_txt = parse_pages_with_vision(tmp_path, target_pages)
                                _os.remove(tmp_path)
                                content_label = f"[문서 첨부: {up_doc.name} (p.{','.join(map(str, target_pages))})]\n{doc_txt[:12000]}"
                            else:
                                with st.spinner("🖼️ VLLM 이미지 분석 중..."):
                                    doc_txt = parse_images_with_vision([(up_doc.name, file_bytes)])
                                content_label = f"[이미지 첨부: {up_doc.name}]\n{doc_txt[:8000]}"
                            msg["doc_handled"] = True
                            chat["messages"].append({"role": "user", "content": content_label})
                            chat["_pending"] = True
                            st.rerun()
                        with col_skip:
                            if st.button("건너뛰기", key=f"idea_doc_skip_{active}_{i}", use_container_width=True):
                                msg["doc_handled"] = True
                                st.rerun()

            # 자동 압축 상태 표시
            if chat.get("compact_summary"):
                with st.expander(f"🗜️ 이전 대화 자동 요약 (LLM이 보는 컨텍스트)", expanded=False):
                    st.markdown(chat["compact_summary"])

            # Pending 상태: 유저 메시지는 이미 표시됨, AI 응답 생성 중
            if chat.get("_pending"):
                with st.chat_message("assistant", avatar="🤖"):
                    with st.status("처리 중...", expanded=True) as status:
                        stage_box = st.empty()
                        text_box = st.empty()
                        search_log_box = st.empty()

                        accumulated_text = ""
                        searches_done = []
                        result_data = {}

                        for ev_type, payload in agentic_chat_stream(
                            chat["messages"], active, linked_objs,
                            overview_text_ideas, "", chat.get("use_search", True),
                            cached_summary=chat.get("compact_summary", ""),
                            cached_until_idx=chat.get("compact_until_idx", 0),
                        ):
                            if ev_type == "stage":
                                status.update(label=payload)
                                stage_box.caption(payload)
                            elif ev_type == "search":
                                searches_done.append(payload)
                                with search_log_box.container():
                                    for q in searches_done:
                                        st.caption(f"🔍 검색: `{q}`")
                            elif ev_type == "text_delta":
                                accumulated_text += payload
                                text_box.markdown(accumulated_text)
                            elif ev_type == "compact":
                                summary_text, until_idx = payload
                                chat["compact_summary"] = summary_text
                                chat["compact_until_idx"] = until_idx
                            elif ev_type == "done":
                                result_data = payload

                        status.update(label="✅ 완료", state="complete", expanded=False)

                # 최종 메시지 저장
                chat["messages"].append({
                    "role": "assistant",
                    "content": result_data.get("text", ""),
                    "search_calls": result_data.get("search_calls", []),
                    "citations": result_data.get("citations", []),
                    "need_doc": result_data.get("need_doc", ""),
                })
                chat["_pending"] = False
                st.rerun()

            # 채팅 입력 + 파일 첨부
            user_input = st.text_area(
                "메시지 입력", placeholder=f"'{active}' 주제로 이야기해주세요...",
                key=f"idea_input_{active}", label_visibility="collapsed", height=80,
            )
            col_file, col_pages, col_send = st.columns([3, 2, 1])
            with col_file:
                attach_file = st.file_uploader(
                    "📎 파일 첨부 (PDF, 이미지)", type=["pdf", "jpg", "jpeg", "png", "gif", "webp", "bmp"],
                    key=f"idea_attach_{active}_{len(chat['messages'])}",
                    label_visibility="collapsed",
                )
            with col_pages:
                attach_pages = st.text_input(
                    "페이지", placeholder="예: 1-5, 10 (PDF만, 비우면 전체)",
                    key=f"idea_pages_{active}_{len(chat['messages'])}",
                    label_visibility="collapsed",
                    disabled=not (attach_file and attach_file.name.lower().endswith(".pdf")),
                )
            with col_send:
                send_clicked = st.button("전송", use_container_width=True, type="primary",
                                         key=f"idea_send_{active}")
            if send_clicked and (user_input.strip() or attach_file):
                from modules.parser import parse_pages_with_vision, parse_images_with_vision
                content_parts = []
                if attach_file:
                    fname = attach_file.name.lower()
                    file_bytes = attach_file.read()
                    if fname.endswith(".pdf"):
                        import tempfile, os as _os
                        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                            tmp.write(file_bytes)
                            tmp_path = tmp.name
                        # 페이지 파싱
                        target_pages = []
                        if attach_pages.strip():
                            for part in attach_pages.split(","):
                                part = part.strip()
                                if "-" in part:
                                    s, e = part.split("-", 1)
                                    try: target_pages.extend(range(int(s), int(e) + 1))
                                    except ValueError: pass
                                else:
                                    try: target_pages.append(int(part))
                                    except ValueError: pass
                        if not target_pages:
                            import fitz
                            doc_f = fitz.open(tmp_path)
                            total_p = len(doc_f)
                            doc_f.close()
                            target_pages = list(range(1, min(total_p + 1, 11)))  # 최대 10페이지
                        with st.spinner(f"📄 VLLM 문서 분석 중... ({len(target_pages)}페이지)"):
                            doc_txt = parse_pages_with_vision(tmp_path, target_pages)
                        _os.remove(tmp_path)
                        content_parts.append(f"[문서 첨부: {attach_file.name} (p.{','.join(map(str, target_pages))})]\n{doc_txt[:12000]}")
                    else:
                        # 이미지 파일
                        with st.spinner("🖼️ VLLM 이미지 분석 중..."):
                            doc_txt = parse_images_with_vision([(attach_file.name, file_bytes)])
                        content_parts.append(f"[이미지 첨부: {attach_file.name}]\n{doc_txt[:8000]}")
                if user_input.strip():
                    content_parts.append(user_input.strip())
                chat["messages"].append({"role": "user", "content": "\n\n".join(content_parts)})
                chat["_pending"] = True
                st.rerun()

            st.divider()

            # 아이디어 정리
            col_sum, col_dl = st.columns(2)
            with col_sum:
                if st.button("🎯 아이디어 정리", use_container_width=True, type="primary",
                             disabled=len(chat["messages"]) < 3):
                    with st.spinner("정리 중..."):
                        summary = summarize_idea_chat(chat["messages"], active, overview_text_ideas)
                    chat["summary"] = summary
                    st.rerun()
            with col_dl:
                if chat.get("summary"):
                    st.download_button(
                        "⬇️ 정리본 다운로드",
                        data=chat["summary"].encode("utf-8"),
                        file_name=f"아이디어_{active}.md",
                        mime="text/markdown",
                        use_container_width=True,
                    )

            # 정리본 표시 (편집 가능)
            if chat.get("summary"):
                with st.expander("📋 정리된 아이디어 (편집 가능)", expanded=True):
                    edited_sum = st.text_area(
                        "정리본",
                        value=chat["summary"],
                        height=400,
                        key=f"sum_edit_{active}",
                        label_visibility="collapsed",
                    )
                    if edited_sum != chat["summary"]:
                        chat["summary"] = edited_sum


# ── 전략 수립 탭 ───────────────────────────────────────────
with tab_strategy:
    from modules.strategist import (
        find_relevant_pages, build_chapter_context,
        parse_agent_response, stream_strategist, stream_critic, run_summarizer,
        check_needed_docs, revise_final_strategy,
    )
    from modules.parser import extract_text_by_page

    debate = st.session_state.debate
    from modules.parser import format_overview_text
    overview_text = format_overview_text(st.session_state.get("project_overview", {}))
    chapters = sorted(set(s.get("chapter", "") for s in st.session_state.slides if s.get("chapter")))

    col_setup, col_debate = st.columns([1, 2])

    with col_setup:
        st.subheader("설정")

        selected_ch = st.selectbox(
            "챕터 선택",
            options=chapters,
            index=chapters.index(debate["chapter"]) if debate["chapter"] in chapters else 0,
            key="debate_chapter_select",
        )

        # 챕터 하위 목차 표시
        ch_slides = [s for s in st.session_state.slides if s.get("chapter") == selected_ch]
        if ch_slides:
            with st.expander(f"하위 목차 ({len(ch_slides)}개)", expanded=True):
                for s in ch_slides:
                    st.caption(f"{s.get('section', '')}. {s.get('title', '')}")

        # 챕터 내 요구사항 체크박스
        ch_req_ids = []
        for s in ch_slides:
            ch_req_ids.extend(s.get("linked_reqs", []))
        ch_req_ids = list(dict.fromkeys(ch_req_ids))

        req_map = {r["id"]: r for r in st.session_state.requirements}
        all_req_ids = [r["id"] for r in st.session_state.requirements]

        fmt_req = lambda rid: f"{rid}  {req_map.get(rid, {}).get('name', '')[:25]}"

        if ch_req_ids:
            # 챕터 매핑 있음: 체크박스 + 추가 선택
            with st.expander(f"요구사항 선택 ({len(ch_req_ids)}개)", expanded=False):
                selected_reqs = []
                for rid in ch_req_ids:
                    r = req_map.get(rid)
                    label = f"`{rid}` {r['name'][:30] if r else rid}"
                    if st.checkbox(label, value=True, key=f"req_check_{rid}"):
                        selected_reqs.append(rid)
            extra_reqs = st.multiselect(
                "추가 요구사항",
                options=[r for r in all_req_ids if r not in ch_req_ids],
                default=[],
                format_func=fmt_req,
                key="req_extra",
                placeholder="챕터 외 요구사항 추가...",
            )
            selected_reqs = selected_reqs + extra_reqs
        else:
            # 매핑 없음: 빈 상태에서 직접 선택
            selected_reqs = st.multiselect(
                f"요구사항 선택 (전체 {len(all_req_ids)}개)",
                options=all_req_ids,
                default=[],
                format_func=fmt_req,
                key="req_multiselect",
                placeholder="요구사항을 선택하세요...",
            )

        max_rounds = st.radio("라운드 수", [2, 3], index=0, horizontal=True)

        # 💡 활용할 아이디어 (정리본이 있는 주제만)
        available_ideas = {
            tname: chat for tname, chat in st.session_state.idea_chats.items()
            if chat.get("summary", "").strip()
        }
        selected_ideas = []
        if available_ideas:
            selected_ideas = st.multiselect(
                "💡 활용할 아이디어",
                options=list(available_ideas.keys()),
                default=[],
                help="기능 아이디어 도출 탭에서 정리한 아이디어를 토론 컨텍스트로 활용",
                key="strategy_ideas_select",
            )
        elif st.session_state.idea_chats:
            st.caption("💡 정리된 아이디어가 없습니다. 아이디어 도출 탭에서 '🎯 아이디어 정리' 먼저 진행")

        if st.button("🎯 전략 수립 시작", use_container_width=True, type="primary"):
            # 선택한 아이디어를 ref_text_cache에 미리 넣기
            initial_ref = ""
            if selected_ideas:
                idea_parts = []
                for tname in selected_ideas:
                    summary = available_ideas[tname].get("summary", "")
                    idea_parts.append(f"=== 💡 아이디어: {tname} ===\n{summary}")
                initial_ref = "\n\n".join(idea_parts)

            st.session_state.debate = {
                "chapter": selected_ch,
                "selected_reqs": selected_reqs,
                "status": "initial_input",
                "history": [],
                "round": 1,
                "max_rounds": max_rounds,
                "pending_question": "",
                "pending_agent": "",
                "pending_doc_request": "",
                "pending_doc_queue": [],
                "ref_text_cache": initial_ref,
                "final_strategy": "",
                "active_ideas": selected_ideas,
                "user_addition": "",
            }
            st.rerun()

        if debate["status"] != "idle":
            st.divider()
            st.caption(f"챕터: **{debate['chapter']}**  |  라운드: {debate['round']}/{debate.get('max_rounds', 2)}")
            if st.button("↩️ 처음부터 다시", use_container_width=True):
                st.session_state.debate["status"] = "idle"
                st.rerun()

    with col_debate:
        st.subheader("장표 작성")

        if debate["status"] == "idle":
            st.info("왼쪽에서 챕터를 선택하고 장표 작성을 시작하세요.")
        else:
            ch = debate["chapter"]
            filtered_reqs = [r for r in st.session_state.requirements
                             if r["id"] in debate.get("selected_reqs", [])]
            if not filtered_reqs:
                filtered_reqs = st.session_state.requirements
            context = build_chapter_context(ch, st.session_state.slides, filtered_reqs)
            ref_text = debate.get("ref_text_cache", "")

            AGENT_LABEL = {"strategist": "🎯 전략가", "critic": "⚔️ 비평가", "user": "👤 답변", "doc": "📎 첨부 문서", "final": "📋 최종 전략"}

            def _next_status(agent):
                is_last = debate["round"] >= debate.get("max_rounds", 2)
                if agent == "strategist":
                    return "critic"
                return "summarizing" if is_last else "strategist"

            # 완료된 턴: expander (최종 전략은 done에서 별도 표시하므로 제외)
            for h in debate["history"]:
                if h["agent"] == "final":
                    continue
                label = f"{AGENT_LABEL.get(h['agent'], h['agent'])}  —  {h['label']}"
                preview = h["content"][:80].replace("\n", " ")
                with st.expander(f"{label}  ·  {preview}...", expanded=False):
                    st.markdown(h["content"])

            status = debate["status"]

            if status == "initial_input":
                st.info("**토론 시작 전, 참고자료와 전략 방향을 미리 알려주세요.** (생략 가능)")
                init_text = st.text_area(
                    "전략 방향 / 사전 의견",
                    height=120,
                    placeholder="예: 우리 회사는 클라우드 보안 인증 보유, 공공 AI 구축 사례 3건. 차별화 포인트로 강조하고 싶음.",
                    key="initial_input_text",
                )
                init_doc = st.file_uploader(
                    "참고 PDF 첨부 (회사소개서, 이전 제안서, 기술백서 등)",
                    type=["pdf"],
                    key=f"initial_doc_{len(debate['history'])}",
                )
                init_pages = ""
                if init_doc:
                    init_pages = st.text_input(
                        "참고할 페이지 (예: 3, 5, 7-9) — 비우면 자동 탐색(텍스트)",
                        key="initial_pages",
                    )
                col_init_doc, col_init_start = st.columns([1, 2])
                with col_init_doc:
                    if init_doc and st.button("문서 반영", use_container_width=True, key="init_add_doc"):
                        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                            tmp.write(init_doc.read())
                            tmp_path = tmp.name
                        pages_sel = parse_page_range(init_pages) if init_pages.strip() else []
                        if pages_sel:
                            doc_text = _vllm_with_progress(
                                lambda cb: parse_pages_with_vision(tmp_path, pages_sel, progress_callback=cb),
                                label=f"VLLM 파싱 ({len(pages_sel)}p)",
                            )
                            label_suffix = f"VLLM p.{','.join(map(str, pages_sel))}"
                        else:
                            with st.spinner("문서 파싱 중..."):
                                doc_pages = extract_text_by_page(tmp_path)
                            relevant = find_relevant_pages(doc_pages, context[:300] + "\n" + init_text[:200])
                            doc_text = "\n\n".join(f"[p{p['page']}]\n{p['text']}" for p in (relevant or doc_pages)[:6])
                            label_suffix = f"자동 {len(relevant or doc_pages)}페이지"
                        os.remove(tmp_path)
                        debate["ref_text_cache"] = (debate.get("ref_text_cache", "") + "\n\n" + doc_text).strip()
                        debate["history"].append({
                            "agent": "doc",
                            "label": f"{init_doc.name} ({label_suffix})",
                            "content": f"문서 첨부됨: **{init_doc.name}** — {label_suffix}",
                        })
                        st.rerun()
                with col_init_start:
                    if st.button("➡️ 토론 시작", use_container_width=True, type="primary", key="init_start"):
                        if init_text.strip():
                            debate["history"].append({
                                "agent": "user", "label": "사전 의견",
                                "content": init_text.strip(),
                            })
                        debate["status"] = "strategist"
                        st.rerun()

            elif status == "strategist":
                with st.chat_message("assistant", avatar="🎯"):
                    st.caption(f"전략가 · 라운드 {debate['round']}")
                    response = st.write_stream(stream_strategist(context, debate["history"], ref_text, overview_text))
                clean, need_info, need_doc = parse_agent_response(response)
                debate["history"].append({"agent": "strategist", "label": f"라운드 {debate['round']}", "content": clean})
                if need_doc:
                    debate["pending_doc_request"] = need_doc
                    debate["pending_agent"] = "strategist"
                    debate["status"] = "waiting_doc"
                elif need_info:
                    debate["pending_question"] = need_info
                    debate["pending_agent"] = "strategist"
                    debate["status"] = "waiting_input"
                else:
                    debate["status"] = "critic"
                st.rerun()

            elif status == "critic":
                with st.chat_message("assistant", avatar="⚔️"):
                    st.caption(f"비평가 · 라운드 {debate['round']}")
                    response = st.write_stream(stream_critic(context, debate["history"], ref_text, overview_text))
                clean, need_info, need_doc = parse_agent_response(response)
                debate["history"].append({"agent": "critic", "label": f"라운드 {debate['round']}", "content": clean})
                if need_doc:
                    debate["pending_doc_request"] = need_doc
                    debate["pending_agent"] = "critic"
                    debate["status"] = "waiting_doc"
                elif need_info:
                    debate["pending_question"] = need_info
                    debate["pending_agent"] = "critic"
                    debate["status"] = "waiting_input"
                elif debate["round"] >= debate.get("max_rounds", 2):
                    debate["status"] = "user_review"
                else:
                    debate["round"] += 1
                    debate["status"] = "strategist"
                st.rerun()

            elif status == "waiting_input":
                agent_name = "전략가" if debate["pending_agent"] == "strategist" else "비평가"
                st.info(f"**{agent_name} 질문:** {debate['pending_question']}")
                user_ans = st.text_area("답변", height=80, label_visibility="collapsed",
                                        placeholder="답변을 입력하세요...", key="debate_ans_input")
                col_ans, col_skip = st.columns([3, 1])
                with col_ans:
                    if st.button("답변 제출", use_container_width=True, type="primary", key="ans_submit"):
                        if user_ans.strip():
                            debate["history"].append({"agent": "user", "label": f"라운드 {debate['round']}", "content": user_ans.strip()})
                        nxt = _next_status(debate["pending_agent"])
                        if nxt == "strategist":
                            debate["round"] += 1
                        debate["status"] = nxt
                        debate["pending_question"] = ""
                        st.rerun()
                with col_skip:
                    if st.button("건너뛰기", use_container_width=True, key="ans_skip"):
                        nxt = _next_status(debate["pending_agent"])
                        if nxt == "strategist":
                            debate["round"] += 1
                        debate["status"] = nxt
                        debate["pending_question"] = ""
                        st.rerun()

            elif status == "waiting_doc":
                pa = debate["pending_agent"]
                queue = debate.get("pending_doc_queue", [])
                queue_info = f" ({len(queue)}건 남음)" if len(queue) > 1 else ""
                if pa == "doc_checker":
                    st.info(f"📎 **장표 작성에 필요한 참고 문서{queue_info}:** {debate['pending_doc_request']}")
                else:
                    agent_name = {"strategist": "전략가", "critic": "비평가"}.get(pa, "AI")
                    st.info(f"**{agent_name}가 문서를 요청했습니다{queue_info}:** {debate['pending_doc_request']}")
                doc_upload = st.file_uploader("PDF 첨부", type=["pdf"], key=f"doc_upload_{len(debate['history'])}")
                wd_pages = ""
                if doc_upload:
                    wd_pages = st.text_input(
                        "참고할 페이지 (예: 3, 5, 7-9) — 비우면 자동 탐색(텍스트)",
                        key=f"doc_pages_{len(debate['history'])}",
                    )

                def _advance_after_doc():
                    q = debate.get("pending_doc_queue", [])
                    if q:
                        q = q[1:]  # 처리된 것 제거
                        debate["pending_doc_queue"] = q
                    if q:
                        nxt_doc = q[0]
                        debate["pending_doc_request"] = f"{nxt_doc.get('name','')} — {nxt_doc.get('reason','')}"
                        debate["status"] = "waiting_doc"
                    else:
                        debate["pending_doc_request"] = ""
                        if pa == "doc_checker":
                            debate["status"] = "done"
                        else:
                            nxt = _next_status(pa)
                            if nxt == "strategist":
                                debate["round"] += 1
                            debate["status"] = nxt

                col_doc, col_skip2 = st.columns([3, 1])
                with col_doc:
                    if doc_upload and st.button("문서 제출", use_container_width=True, type="primary"):
                        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                            tmp.write(doc_upload.read())
                            tmp_path = tmp.name
                        pages_sel = parse_page_range(wd_pages) if wd_pages.strip() else []
                        if pages_sel:
                            doc_text = _vllm_with_progress(
                                lambda cb: parse_pages_with_vision(tmp_path, pages_sel, progress_callback=cb),
                                label=f"VLLM 파싱 ({len(pages_sel)}p)",
                            )
                            label_suffix = f"VLLM p.{','.join(map(str, pages_sel))}"
                        else:
                            with st.spinner("문서 파싱 중..."):
                                doc_pages = extract_text_by_page(tmp_path)
                            relevant = find_relevant_pages(doc_pages, context[:300] + "\n" + debate["pending_doc_request"][:200])
                            doc_text = "\n\n".join(f"[p{p['page']}]\n{p['text']}" for p in (relevant or doc_pages)[:6])
                            label_suffix = f"자동 {len(relevant or doc_pages)}페이지"
                        os.remove(tmp_path)
                        debate["ref_text_cache"] = (debate.get("ref_text_cache", "") + "\n\n" + doc_text).strip()
                        debate["history"].append({
                            "agent": "doc",
                            "label": f"{doc_upload.name} ({label_suffix})",
                            "content": f"문서 첨부됨: **{doc_upload.name}** — {label_suffix}",
                        })
                        _advance_after_doc()
                        st.rerun()
                with col_skip2:
                    if st.button("건너뛰기", use_container_width=True):
                        _advance_after_doc()
                        st.rerun()

            elif status == "user_review":
                st.info("**토론이 끝났습니다.** 정리 단계로 넘어가기 전에 추가 의견이나 참고문서를 넣을 수 있습니다.")
                ur_text = st.text_area(
                    "추가 의견",
                    height=100,
                    placeholder="예: PoC 경험 강조, 보안 영역 보완, 특정 솔루션명 사용 등",
                    key="user_review_text",
                )
                ur_doc = st.file_uploader(
                    "참고 PDF 첨부", type=["pdf"],
                    key=f"user_review_doc_{len(debate['history'])}",
                )
                ur_pages = ""
                if ur_doc:
                    ur_pages = st.text_input(
                        "참고할 페이지 (예: 3, 5, 7-9) — 비우면 자동 탐색(텍스트 추출)",
                        key="user_review_pages",
                    )
                col_doc_ur, col_proceed = st.columns([1, 2])
                with col_doc_ur:
                    if ur_doc and st.button("문서 반영", use_container_width=True, key="ur_add_doc"):
                        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                            tmp.write(ur_doc.read())
                            tmp_path = tmp.name
                        pages = parse_page_range(ur_pages) if ur_pages.strip() else []
                        if pages:
                            doc_text = _vllm_with_progress(
                                lambda cb: parse_pages_with_vision(tmp_path, pages, progress_callback=cb),
                                label=f"VLLM 파싱 ({len(pages)}p)",
                            )
                            label_suffix = f"VLLM p.{','.join(map(str, pages))}"
                        else:
                            with st.spinner("문서 파싱 중..."):
                                doc_pages = extract_text_by_page(tmp_path)
                            relevant = find_relevant_pages(doc_pages, context[:300] + "\n" + ur_text[:200])
                            doc_text = "\n\n".join(f"[p{p['page']}]\n{p['text']}" for p in (relevant or doc_pages)[:6])
                            label_suffix = f"자동 {len(relevant or doc_pages)}페이지"
                        os.remove(tmp_path)
                        debate["ref_text_cache"] = (debate.get("ref_text_cache", "") + "\n\n" + doc_text).strip()
                        debate["history"].append({
                            "agent": "doc",
                            "label": f"{ur_doc.name} ({label_suffix})",
                            "content": f"문서 첨부됨: **{ur_doc.name}** — {label_suffix}",
                        })
                        st.rerun()
                with col_proceed:
                    if st.button("➡️ 전략 정리로 진행", use_container_width=True, type="primary", key="ur_proceed"):
                        if ur_text.strip():
                            debate["history"].append({
                                "agent": "user", "label": "최종 의견",
                                "content": ur_text.strip(),
                            })
                        debate["status"] = "summarizing"
                        st.rerun()

            elif status == "summarizing":
                with st.spinner("최종 전략 정리 중..."):
                    final = run_summarizer(context, debate["history"], overview_text)
                debate["final_strategy"] = final
                debate["history"].append({"agent": "final", "label": "최종", "content": final})
                debate["status"] = "checking_docs"
                st.rerun()

            elif status == "checking_docs":
                with st.spinner("장표 작성에 필요한 참고문서 확인 중..."):
                    needed = check_needed_docs(context, debate["final_strategy"], ref_text, overview_text)
                if needed:
                    doc_list_text = "**장표 작성을 위해 다음 문서가 필요합니다:**\n\n" + "\n".join(
                        f"- **{d.get('name','')}**: {d.get('reason','')}" for d in needed
                    )
                    debate["history"].append({
                        "agent": "doc_check", "label": "문서 체크", "content": doc_list_text
                    })
                    first = needed[0]
                    debate["pending_doc_queue"] = needed
                    debate["pending_doc_request"] = f"{first.get('name','')} — {first.get('reason','')}"
                    debate["pending_agent"] = "doc_checker"
                    debate["status"] = "waiting_doc"
                else:
                    debate["status"] = "done"
                st.rerun()

            elif status == "done":
                with st.expander("📋 최종 전략", expanded=True):
                    st.markdown(debate["final_strategy"])

                if ref_text.strip():
                    with st.expander(f"📎 참고 문서 ({len(ref_text):,}자)", expanded=False):
                        st.text(ref_text[:5000] + ("\n\n... (이하 생략)" if len(ref_text) > 5000 else ""))

                st.divider()
                st.markdown("**장표별 초안 작성**")
                ch_slides = [s for s in st.session_state.slides if s.get("chapter") == ch]
                strategy_hint = debate["final_strategy"]

                # 챕터 전체 outline 일괄 생성 (서로 침범 방지)
                if ch_slides:
                    missing_outline_count = sum(1 for s in ch_slides if not s.get("outline"))
                    if st.button(
                        f"🎯 챕터 전체 outline 일괄 생성 (서로 침범 방지)  [{ch}장 {len(ch_slides)}개 슬라이드]",
                        use_container_width=True,
                        type="secondary",
                        help="모든 슬라이드의 outline을 한 번에 생성하여 서로 다루는 영역이 겹치지 않도록 자동 분배합니다. 기존 outline은 덮어씁니다.",
                    ):
                        with st.spinner(f"{len(ch_slides)}개 슬라이드 outline 일괄 생성 중..."):
                            result = generate_chapter_outlines(
                                ch_slides, st.session_state.requirements,
                                strategy_hint, ref_text, overview_text,
                            )
                        if not result:
                            st.error("일괄 생성 실패. 개별 생성을 시도하거나 다시 시도하세요.")
                        else:
                            applied = 0
                            for slide in ch_slides:
                                sid = slide.get("id", "")
                                if sid in result:
                                    slide["outline"] = result[sid]
                                    # 기존 sections는 그대로 유지하되, outline이 바뀐 항목들은 비움
                                    new_titles = {o["title"] for o in result[sid]}
                                    if slide.get("sections"):
                                        slide["sections"] = {
                                            k: v for k, v in slide["sections"].items() if k in new_titles
                                        }
                                    # 편집 위젯 값 동기화
                                    s_key = slide.get("id", slide.get("title", ""))
                                    for t in new_titles:
                                        ek = f"sec_edit_{s_key}_{t}"
                                        if t in slide.get("sections", {}):
                                            st.session_state[ek] = slide["sections"][t]
                                        else:
                                            st.session_state.pop(ek, None)
                                    applied += 1
                            st.success(f"✅ {applied}/{len(ch_slides)}개 슬라이드 outline 생성 완료")
                            st.rerun()
                for slide in ch_slides:
                    s_key = slide.get("id", slide.get("title", ""))
                    slide_reqs = [r for r in st.session_state.requirements if r["id"] in slide.get("linked_reqs", [])]
                    icon = "✅" if slide.get("status") == "초안완료" else "⬜"
                    with st.expander(f"{icon} {slide.get('section', '')}. {slide.get('title', '')}", expanded=False):
                        # 목차 생성
                        outline = slide.get("outline", [])
                        col_ol, col_re = st.columns([3, 1])
                        with col_ol:
                            if not outline:
                                st.caption("목차를 생성하면 소제목별로 내용을 작성합니다.")
                        with col_re:
                            btn_label = "목차 재생성" if outline else "목차 생성"
                            if st.button(btn_label, key=f"outline_{s_key}", use_container_width=True):
                                with st.spinner("목차 생성 중..."):
                                    siblings = [s for s in ch_slides if s.get("id") != slide.get("id")]
                                    slide["outline"] = generate_outline(
                                        slide, slide_reqs, strategy_hint, ref_text, overview_text,
                                        sibling_slides=siblings,
                                    )
                                    slide.setdefault("sections", {})
                                st.rerun()

                        if outline:
                            # 구버전 호환: 문자열 리스트면 dict 리스트로 변환
                            outline = [
                                {"title": o, "scope": ""} if isinstance(o, str) else o
                                for o in outline
                            ]
                            sections = slide.get("sections", {})
                            titles = [o["title"] for o in outline]
                            missing = [o for o in outline if o["title"] not in sections]

                            # 일괄 생성 (병렬)
                            if missing:
                                if st.button(
                                    f"⚡ 미작성 소제목 병렬 일괄 생성 ({len(missing)}개)",
                                    key=f"bulk_{s_key}",
                                    use_container_width=True,
                                    type="primary",
                                ):
                                    from concurrent.futures import ThreadPoolExecutor

                                    def _run(o):
                                        return o["title"], generate_section(
                                            slide, o["title"], o.get("scope", ""),
                                            slide_reqs, strategy_hint, ref_text, overview_text,
                                        )
                                    with st.spinner(f"{len(missing)}개 병렬 생성 중..."):
                                        with ThreadPoolExecutor(max_workers=min(len(missing), 4)) as ex:
                                            results = list(ex.map(_run, missing))
                                    slide.setdefault("sections", {})
                                    for t, content in results:
                                        slide["sections"][t] = content
                                        # 편집 위젯 값도 동기화
                                        st.session_state[f"sec_edit_{s_key}_{t}"] = content
                                    slide["draft"] = "\n\n".join(
                                        f"### {t}\n{slide['sections'][t]}" for t in titles
                                        if t in slide["sections"]
                                    )
                                    if all(t in slide["sections"] for t in titles):
                                        slide["status"] = "초안완료"
                                    st.rerun()

                            for o in outline:
                                sec_title = o["title"]
                                sec_scope = o.get("scope", "")
                                has_content = sec_title in sections
                                edit_key = f"sec_edit_{s_key}_{sec_title}"

                                # 헤더 + 액션 버튼
                                col_h, col_btn1, col_btn2 = st.columns([4, 1, 1])
                                with col_h:
                                    icon = "✏️" if has_content else "⬜"
                                    st.markdown(f"{icon} **{sec_title}**")
                                    if sec_scope:
                                        st.caption(f"📐 scope: {sec_scope}")
                                with col_btn1:
                                    btn_label = "재생성" if has_content else "AI 작성"
                                    if st.button(btn_label, key=f"sec_btn_{s_key}_{sec_title}", use_container_width=True):
                                        with st.spinner(f"'{sec_title}' 작성 중..."):
                                            content = st.write_stream(
                                                generate_section_stream(slide, sec_title, slide_reqs, strategy_hint, sec_scope, ref_text, overview_text)
                                            )
                                        slide.setdefault("sections", {})[sec_title] = content
                                        st.session_state[edit_key] = content
                                        if all(t in slide.get("sections", {}) for t in titles):
                                            slide["draft"] = "\n\n".join(
                                                f"### {t}\n{slide['sections'][t]}" for t in titles
                                            )
                                            slide["status"] = "초안완료"
                                        st.rerun()
                                with col_btn2:
                                    if has_content and st.button("지우기", key=f"sec_clear_{s_key}_{sec_title}", use_container_width=True):
                                        slide.get("sections", {}).pop(sec_title, None)
                                        st.session_state.pop(edit_key, None)
                                        st.rerun()

                                # 편집 가능한 본문
                                if edit_key not in st.session_state:
                                    st.session_state[edit_key] = sections.get(sec_title, "")
                                edited = st.text_area(
                                    f"본문_{sec_title}",
                                    key=edit_key,
                                    height=180,
                                    label_visibility="collapsed",
                                    placeholder="여기에 직접 작성하거나 'AI 작성' 버튼을 누르세요.",
                                )
                                # 편집 결과를 슬라이드에 반영
                                stored = sections.get(sec_title, "")
                                if edited != stored:
                                    if edited.strip():
                                        slide.setdefault("sections", {})[sec_title] = edited
                                    else:
                                        slide.get("sections", {}).pop(sec_title, None)
                                    # 초안 재조립
                                    if all(t in slide.get("sections", {}) for t in titles):
                                        slide["draft"] = "\n\n".join(
                                            f"### {t}\n{slide['sections'][t]}" for t in titles
                                        )
                                        slide["status"] = "초안완료"
                                    elif slide.get("sections"):
                                        slide["draft"] = "\n\n".join(
                                            f"### {t}\n{slide['sections'][t]}" for t in titles if t in slide["sections"]
                                        )

                            # 수동 조립
                            written = [t for t in titles if t in slide.get("sections", {})]
                            if written and len(written) < len(titles):
                                if st.button("지금까지 내용 조립", key=f"assemble_{s_key}", use_container_width=True):
                                    slide["draft"] = "\n\n".join(
                                        f"### {t}\n{slide['sections'][t]}" for t in titles if t in slide["sections"]
                                    )
                                    slide["status"] = "초안완료"
                                    st.rerun()

                # 챕터 마크다운 다운로드
                completed = [s for s in ch_slides if s.get("draft")]
                if completed:
                    st.divider()
                    AGENT_MD = {
                        "strategist": "🎯 전략가",
                        "critic": "⚔️ 비평가",
                        "user": "👤 사용자",
                        "doc": "📎 첨부 문서",
                        "doc_check": "🔍 문서 체크",
                        "final": "📋 최종 전략",
                    }

                    md_lines = [
                        f"# {ch}장 — 제안서 초안 패키지\n",
                        "> 이 문서는 RFP 분석 → 전략 토론 → 참고문서 반영 → 장표별 초안 작성 전체 컨텍스트를 담고 있습니다.\n",
                        "---\n",
                    ]
                    if overview_text:
                        md_lines.extend([
                            "## 🏢 사업 개요\n",
                            overview_text,
                            "\n---\n",
                        ])
                    md_lines.extend([
                        "## 📑 챕터 구성 및 요구사항\n",
                        "```\n" + context + "\n```\n",
                        "---\n",
                        "## 📋 최종 전략\n",
                        debate["final_strategy"],
                        "\n---\n",
                    ])

                    # 전략 논의 history
                    hist_items = [h for h in debate["history"] if h["agent"] not in ("final",)]
                    if hist_items:
                        md_lines.append("## 💬 전략 논의 (History)\n")
                        for h in hist_items:
                            md_lines.append(f"### {AGENT_MD.get(h['agent'], h['agent'])} — {h['label']}\n")
                            md_lines.append(h["content"])
                            md_lines.append("\n")
                        md_lines.append("---\n")

                    # 참고문서 (VLLM/텍스트 추출)
                    if ref_text.strip():
                        md_lines.append("## 📎 참고 문서 (추출 데이터)\n")
                        md_lines.append(ref_text)
                        md_lines.append("\n---\n")

                    # 장표별 초안
                    md_lines.append("## ✍️ 장표별 초안\n")
                    for slide in ch_slides:
                        if not slide.get("draft"):
                            continue
                        md_lines.append(f"### {slide.get('section','')}. {slide.get('title','')}\n")
                        slide_reqs = [r for r in st.session_state.requirements if r["id"] in slide.get("linked_reqs", [])]
                        if slide_reqs:
                            md_lines.append("**연결 요구사항:**\n")
                            for r in slide_reqs:
                                name = r.get("name", "").strip() or "(이름 없음)"
                                md_lines.append(f"- **`{r.get('id','')}` {name}**")
                                md_lines.append(f"  - 분류: {r.get('category','') or '(미분류)'}")
                                # definition 비어있으면 detail 첫 문장으로 fallback
                                defi = (r.get("definition") or "").strip()
                                detail = (r.get("detail") or "").replace("\n", " ").strip()
                                if not defi and detail:
                                    defi = detail.split(".")[0][:120].strip() + "..."
                                md_lines.append(f"  - 정의: {defi or '(없음)'}")
                                md_lines.append(f"  - 세부내용: {detail or '(없음)'}")
                            md_lines.append("")
                        md_lines.append("**초안:**\n")
                        md_lines.append(slide["draft"])
                        md_lines.append("\n")

                    md_content = "\n".join(md_lines)
                    safe_ch = str(ch).replace("/", "_").replace("\\", "_")
                    st.download_button(
                        label=f"📥 {ch}장 마크다운 다운로드 ({len(completed)}/{len(ch_slides)}장표, 전체 컨텍스트 포함)",
                        data=md_content.encode("utf-8"),
                        file_name=f"{safe_ch}장_제안서패키지.md",
                        mime="text/markdown",
                        use_container_width=True,
                        key=f"download_md_{ch}",
                    )
                    st.caption(f"📦 {len(md_content):,}자 — Claude에 그대로 붙여 'PPT 만들어줘' 가능")
